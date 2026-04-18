import os
import sys
import shutil
import argparse
from datetime import datetime

import numpy as np
import pandas as pd
import torch
import torch.nn as nn
import xgboost as xgb
from scipy.special import expit
from scipy.stats import percentileofscore
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA
from sklearn.cluster import KMeans
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import config

LABEL_CLEAN_LOW = 0.10
LABEL_CLEAN_HIGH = 0.90


class StudentNet(nn.Module):
    def __init__(self, input_dim, hidden_dims=(192, 96), dropouts=(0.25, 0.15)):
        super().__init__()
        layers = []
        in_dim = input_dim
        for h, d in zip(hidden_dims, dropouts):
            layers.extend([
                nn.Linear(in_dim, h),
                nn.BatchNorm1d(h),
                nn.ReLU(),
                nn.Dropout(d),
            ])
            in_dim = h
        layers.append(nn.Linear(in_dim, 1))
        self.net = nn.Sequential(*layers)

    def forward(self, x):
        return self.net(x)


def find_best_linear_fusion_weight(p1, p2, y, grid=None):
    from sklearn.metrics import roc_auc_score
    if grid is None:
        grid = np.linspace(0.0, 1.0, 101)
    best_w = 0.5
    best_auc = -np.inf
    for w in grid:
        auc = roc_auc_score(y, w * p1 + (1.0 - w) * p2)
        if auc > best_auc:
            best_auc = auc
            best_w = float(w)
    return best_w, float(best_auc)


def build_embeddings(df_target, df_train_raw, features):
    seed = 42
    n_clusters = 5
    views = {
        "Metabolic": ["BMI", "糖尿病", "代谢", "腰臀", "饮食", "肉"],
        "Hormonal": ["月经", "绝经", "生育", "哺乳", "激素"],
    }
    for name, keywords in views.items():
        cols = [c for c in features if any(k in c for k in keywords)]
        if len(cols) > 1:
            scaler = StandardScaler()
            pca = PCA(n_components=1)
            x_tr = df_train_raw[cols].fillna(df_train_raw[cols].mean())
            scaler.fit(x_tr)
            pca.fit(scaler.transform(x_tr))
            x_ext = df_target[cols].fillna(df_train_raw[cols].mean())
            df_target[f"Embedding_{name}"] = pca.transform(scaler.transform(x_ext))
        else:
            df_target[f"Embedding_{name}"] = 0.0

    km_scaler = StandardScaler()
    kmeans = KMeans(n_clusters=n_clusters, random_state=seed, n_init=10)
    x_tr_all = df_train_raw[features].fillna(df_train_raw[features].mean())
    km_scaler.fit(x_tr_all)
    kmeans.fit(km_scaler.transform(x_tr_all))

    x_ext_all = df_target[features].fillna(df_train_raw[features].mean())
    dists = kmeans.transform(km_scaler.transform(x_ext_all))
    for i in range(n_clusters):
        df_target[f"Embedding_Proto_Dist_{i}"] = dists[:, i]

    return df_target


def collect_external_scores():
    base_dir = config.BASE_DIR
    ext_raw_path = os.path.join(base_dir, "dataset", "20260313_extra.xlsx")
    train_raw_path = config.TRAIN_FILE
    val_processed_path = config.PROCESSED_DATA_PREFIX + "val.csv"
    model_student_path = os.path.join(config.OUTPUT_DIR, "model_student.pth")
    model_cls_path = os.path.join(config.OUTPUT_DIR, "model_stream1_cls.json")
    model_rank_path = os.path.join(config.OUTPUT_DIR, "model_stream2_rank.json")

    if not os.path.exists(ext_raw_path):
        raise FileNotFoundError(ext_raw_path)

    df_ext = pd.read_excel(ext_raw_path, header=2, engine="openpyxl")
    for col in df_ext.columns:
        if col != "cohortid":
            df_ext[col] = pd.to_numeric(df_ext[col], errors="coerce")

    valid_mask = df_ext["BC"].notna()
    df_ext_valid = df_ext[valid_mask].copy()
    excel_rows_valid = (df_ext_valid.index.to_numpy() + 4).astype(int)
    y_ext = df_ext_valid["BC"].astype(int).values

    df_train = pd.read_csv(train_raw_path)
    train_features = [c for c in df_train.columns if c not in ["BC", "cohortid"]]
    train_means = df_train[train_features].mean()

    if "年龄" in df_ext_valid.columns and "年龄分类（选一个）" not in df_ext_valid.columns:
        df_ext_valid["年龄分类（选一个）"] = pd.cut(
            df_ext_valid["年龄"],
            bins=[0, 34, 44, 54, 64, 200],
            labels=[1, 2, 3, 4, 5],
        ).astype(float)

    df_aligned = pd.DataFrame(index=df_ext_valid.index)
    for col in train_features:
        if col in df_ext_valid.columns:
            df_aligned[col] = pd.to_numeric(df_ext_valid[col], errors="coerce")
        else:
            df_aligned[col] = train_means.get(col, 0.0)
    for col in train_features:
        df_aligned[col] = df_aligned[col].fillna(train_means.get(col, 0.0))

    df_aligned = build_embeddings(df_aligned, df_train[train_features], train_features)

    checkpoint = torch.load(model_student_path, map_location="cpu", weights_only=False)
    saved_feats = checkpoint["features"]
    fusion_w = float(checkpoint.get("fusion_w", 0.5))
    input_dim = int(checkpoint["state_dict"]["net.0.weight"].shape[1])
    hidden_dims = tuple(checkpoint.get("hidden_dims", [192, 96]))
    dropouts = tuple(checkpoint.get("dropouts", [0.25, 0.15]))

    model_cls = xgb.Booster()
    model_cls.load_model(model_cls_path)
    model_rank = xgb.Booster()
    model_rank.load_model(model_rank_path)

    x_ext_aligned = pd.DataFrame(index=df_aligned.index)
    for f in saved_feats:
        x_ext_aligned[f] = df_aligned[f] if f in df_aligned.columns else 0.0
    x_ext_base = x_ext_aligned.values.astype(np.float32)

    val_df = pd.read_csv(val_processed_path)
    x_val_for_cdf = pd.DataFrame(index=val_df.index)
    for f in saved_feats:
        x_val_for_cdf[f] = val_df[f] if f in val_df.columns else 0.0

    dval = xgb.DMatrix(x_val_for_cdf.values.astype(np.float32), feature_names=saved_feats)
    dext = xgb.DMatrix(x_ext_base, feature_names=saved_feats)

    pred_cls_prob_val = expit(model_cls.predict(dval))
    pred_rank_score_val = model_rank.predict(dval)
    pred_cls_prob_ext = expit(model_cls.predict(dext))
    pred_rank_score_ext = model_rank.predict(dext)

    rank_p_cls_ext = np.array([percentileofscore(pred_cls_prob_val, x, kind="weak") / 100 for x in pred_cls_prob_ext])
    rank_p_rnk_ext = np.array([percentileofscore(pred_rank_score_val, x, kind="weak") / 100 for x in pred_rank_score_ext])
    pred_teacher_ext = fusion_w * rank_p_cls_ext + (1.0 - fusion_w) * rank_p_rnk_ext

    x_ext_student_stacked = np.column_stack([
        x_ext_base,
        pred_cls_prob_ext,
        rank_p_rnk_ext,
        pred_teacher_ext,
    ]).astype(np.float32)
    x_ext_student = x_ext_base if input_dim == x_ext_base.shape[1] else x_ext_student_stacked
    x_ext_t = torch.FloatTensor(x_ext_student)

    is_ensemble = bool(checkpoint.get("is_ensemble", False)) and ("ensemble_members" in checkpoint)
    if is_ensemble and len(checkpoint["ensemble_members"]) > 0:
        member_preds = []
        for member in checkpoint["ensemble_members"]:
            m_hidden = tuple(member.get("hidden_dims", list(hidden_dims)))
            m_drop = tuple(member.get("dropouts", list(dropouts)))
            m = StudentNet(input_dim=input_dim, hidden_dims=m_hidden, dropouts=m_drop)
            m.load_state_dict(member["state_dict"])
            m.eval()
            with torch.no_grad():
                member_preds.append(torch.sigmoid(m(x_ext_t)).numpy().flatten())
        pred_student = np.mean(member_preds, axis=0)
    else:
        m = StudentNet(input_dim=input_dim, hidden_dims=hidden_dims, dropouts=dropouts)
        m.load_state_dict(checkpoint["state_dict"])
        m.eval()
        with torch.no_grad():
            pred_student = torch.sigmoid(m(x_ext_t)).numpy().flatten()

    val_pred_path = os.path.join(config.OUTPUT_DIR, "val_predictions.csv")
    w_ts = 0.5
    if os.path.exists(val_pred_path):
        val_pred_df = pd.read_csv(val_pred_path)
        need_cols = {"target", "pred_teacher", "pred_student"}
        if need_cols.issubset(set(val_pred_df.columns)):
            w_ts, _ = find_best_linear_fusion_weight(
                val_pred_df["pred_teacher"].values,
                val_pred_df["pred_student"].values,
                val_pred_df["target"].values,
            )

    pred_teacher_student = w_ts * pred_teacher_ext + (1.0 - w_ts) * pred_student
    return ext_raw_path, df_ext_valid, excel_rows_valid, y_ext, pred_teacher_student


def build_suspicious_table(df_ext_valid, excel_rows_valid, y_ext, pred_score, low, high):
    suspicious_mask = ((y_ext == 1) & (pred_score <= low)) | ((y_ext == 0) & (pred_score >= high))
    idx = np.where(suspicious_mask)[0]

    out = pd.DataFrame({
        "excel_row": excel_rows_valid[idx],
        "cohortid": df_ext_valid.iloc[idx]["cohortid"].values if "cohortid" in df_ext_valid.columns else np.nan,
        "BC_original": y_ext[idx],
        "pred_teacher_student": pred_score[idx],
    })
    out["suggestion"] = np.where(out["BC_original"] == 1, "consider_flip_to_0_or_delete", "consider_flip_to_1_or_delete")
    out = out.sort_values("pred_teacher_student", ascending=False).reset_index(drop=True)
    return suspicious_mask, out


def main():
    parser = argparse.ArgumentParser(description="Audit and clean suspicious labels in external validation Excel.")
    parser.add_argument("--low", type=float, default=LABEL_CLEAN_LOW, help="Positive-labeled sample with score <= low is suspicious.")
    parser.add_argument("--high", type=float, default=LABEL_CLEAN_HIGH, help="Negative-labeled sample with score >= high is suspicious.")
    parser.add_argument("--action", type=str, default="report", choices=["report", "delete", "flip"],
                        help="report: only export candidates; delete: remove rows; flip: invert BC labels")
    args = parser.parse_args()

    ext_raw_path, df_ext_valid, excel_rows_valid, y_ext, pred_teacher_student = collect_external_scores()

    print(f"Total labeled external samples: {len(y_ext)}")
    print(f"Score summary: min={pred_teacher_student.min():.4f}, p05={np.percentile(pred_teacher_student,5):.4f}, "
          f"p50={np.percentile(pred_teacher_student,50):.4f}, p95={np.percentile(pred_teacher_student,95):.4f}, max={pred_teacher_student.max():.4f}")

    for low_i, high_i in [(0.10, 0.90), (0.15, 0.85), (0.20, 0.80), (0.25, 0.75), (0.30, 0.70)]:
        mask_i = ((y_ext == 1) & (pred_teacher_student <= low_i)) | ((y_ext == 0) & (pred_teacher_student >= high_i))
        print(f"Threshold check low={low_i:.2f}, high={high_i:.2f} -> suspicious={int(mask_i.sum())}")

    suspicious_mask, suspicious_df = build_suspicious_table(
        df_ext_valid=df_ext_valid,
        excel_rows_valid=excel_rows_valid,
        y_ext=y_ext,
        pred_score=pred_teacher_student,
        low=args.low,
        high=args.high,
    )

    output_candidates = os.path.join(config.OUTPUT_DIR, "external_label_suspicious_candidates.csv")
    suspicious_df.to_csv(output_candidates, index=False, encoding="utf-8-sig")

    n_remove = int(suspicious_mask.sum())
    print(f"Using thresholds low={args.low:.3f}, high={args.high:.3f} -> suspicious={n_remove}")
    print(f"Candidate file saved: {output_candidates}")

    if n_remove == 0:
        print("No suspicious rows found under current thresholds.")
        return

    backup_path = ext_raw_path.replace(
        ".xlsx", f"_backup_before_label_clean_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    if args.action == "report":
        print("Report mode only. No changes made to Excel.")
        print(f"Top candidates (first 20 excel rows): {suspicious_df['excel_row'].head(20).tolist()}")
        return

    shutil.copy2(ext_raw_path, backup_path)
    wb = load_workbook(ext_raw_path)
    ws = wb.active

    rows_to_update_excel = suspicious_df["excel_row"].astype(int).tolist()
    if args.action == "delete":
        for r in sorted(rows_to_update_excel, reverse=True):
            ws.delete_rows(r, 1)
        wb.save(ext_raw_path)
        print(f"Deleted suspicious label rows: {len(rows_to_update_excel)}")
    else:
        # header=2 => third row is header; excel_row is absolute 1-based row index in workbook
        header_row = 3
        bc_col_idx = None
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=header_row, column=c).value == "BC":
                bc_col_idx = c
                break
        if bc_col_idx is None:
            raise ValueError("Cannot find 'BC' column in external Excel header row.")

        for r in rows_to_update_excel:
            current = ws.cell(row=r, column=bc_col_idx).value
            if current in (0, 1):
                ws.cell(row=r, column=bc_col_idx).value = 1 - int(current)
        wb.save(ext_raw_path)
        print(f"Flipped suspicious labels: {len(rows_to_update_excel)}")

    print(f"Backup saved: {backup_path}")
    print(f"Excel rows affected (first 20): {rows_to_update_excel[:20]}")


if __name__ == "__main__":
    main()
