# Project path bootstrap for direct script execution from nested folders.
from pathlib import Path
import sys

for _project_root in Path(__file__).resolve().parents:
    if (_project_root / "configs").is_dir() and (_project_root / "utils").is_dir():
        for _path in (_project_root, _project_root / "configs", _project_root / "utils"):
            _path_str = str(_path)
            if _path_str not in sys.path:
                sys.path.insert(0, _path_str)
        break
# 06_ablation_study.py
# 消融实验：输出五个核心结果
# 消融项：
#   1) 仅原始特征 + 单流 XGB
#   2) 全部特征 + 单流 XGB
#   3) 全部特征 + 直接 MLP
#   4) 全部特征 + Teacher（复用 0313）
#   5) 全部特征 + Student（复用 0313）
# 输出: ablation_table.csv, ablation_table.xlsx

import pandas as pd
import numpy as np
import xgboost as xgb
import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import TensorDataset, DataLoader
import copy
from scipy.special import expit
from scipy.stats import rankdata, percentileofscore
from sklearn.metrics import roc_auc_score, roc_curve, confusion_matrix
import config
import os
import warnings

warnings.filterwarnings("ignore")


# ==========================================
# 复用 03_train.py 的核心组件
# ==========================================

def asymmetric_focal_loss_obj(pred, dtrain):
    y_true = dtrain.get_label()
    p = expit(pred)
    gamma = 2.0
    alpha = 4.0
    residual = p - y_true
    modulating_factor = np.where(y_true == 1, (1 - p) ** gamma, p ** gamma)
    alpha_weight = np.where(y_true == 1, alpha, 1.0)
    grad = alpha_weight * modulating_factor * residual + \
           alpha_weight * gamma * (1 - p) * p * modulating_factor * np.log(np.clip(p, 1e-6, 1.0 - 1e-6)) * (
                       2 * y_true - 1)
    hess = alpha_weight * modulating_factor * p * (1 - p)
    hess = np.maximum(hess, 1e-6)
    return grad, hess


class StudentNet(nn.Module):
    def __init__(self, input_dim):
        super().__init__()
        self.net = nn.Sequential(
            nn.Linear(input_dim, 256), nn.BatchNorm1d(256), nn.ReLU(), nn.Dropout(0.4),
            nn.Linear(256, 128), nn.BatchNorm1d(128), nn.ReLU(), nn.Dropout(0.3),
            nn.Linear(128, 64), nn.BatchNorm1d(64), nn.ReLU(), nn.Dropout(0.2),
            nn.Linear(64, 1),
        )

    def forward(self, x):
        return self.net(x)


def make_ranking_data(X, y, group_size=1000):
    perm = np.random.permutation(len(X))
    X_shuff = X.iloc[perm].reset_index(drop=True)
    y_shuff = y.iloc[perm].reset_index(drop=True)
    n_samples = len(X)
    n_groups = int(np.ceil(n_samples / group_size))
    group_counts = [group_size] * (n_groups - 1)
    group_counts.append(n_samples - sum(group_counts))
    return X_shuff, y_shuff, group_counts


def find_best_fusion_weight(p1, p2, y, grid=None):
    if grid is None:
        grid = np.linspace(0.0, 1.0, 101)
    best_w, best_auc = 0.5, -np.inf
    for w in grid:
        auc = roc_auc_score(y, w * p1 + (1.0 - w) * p2)
        if auc > best_auc:
            best_auc, best_w = auc, float(w)
    return best_w, float(best_auc)


def bootstrap_auc_ci(y_true, y_prob, n_boot=1000, seed=42):
    rng = np.random.default_rng(seed)
    n = len(y_true)
    y_true = np.asarray(y_true)
    y_prob = np.asarray(y_prob)
    aucs = []
    for _ in range(n_boot):
        idx = rng.integers(0, n, n)
        if len(np.unique(y_true[idx])) < 2:
            continue
        aucs.append(roc_auc_score(y_true[idx], y_prob[idx]))
    return np.percentile(aucs, [2.5, 97.5])


def youden_threshold(y_true, y_prob):
    fpr, tpr, thresholds = roc_curve(y_true, y_prob)
    j = tpr - fpr
    return float(thresholds[np.argmax(j)])


def clinical_metrics(y_true, y_prob, threshold):
    y_pred = (y_prob >= threshold).astype(int)
    tn, fp, fn, tp = confusion_matrix(y_true, y_pred, labels=[0, 1]).ravel()
    sens = tp / (tp + fn) if (tp + fn) > 0 else 0.0
    spec = tn / (tn + fp) if (tn + fp) > 0 else 0.0
    ppv = tp / (tp + fp) if (tp + fp) > 0 else 0.0
    npv = tn / (tn + fn) if (tn + fn) > 0 else 0.0
    return sens, spec, ppv, npv


def bootstrap_metric_ci(y_true, y_prob, threshold, n_boot=1000, seed=42):
    rng = np.random.default_rng(seed)
    y_true = np.asarray(y_true)
    y_prob = np.asarray(y_prob)
    n = len(y_true)

    sens_list, spec_list, ppv_list, npv_list = [], [], [], []
    for _ in range(n_boot):
        idx = rng.integers(0, n, n)
        y_b = y_true[idx]
        p_b = y_prob[idx]
        s, sp, p, n_ = clinical_metrics(y_b, p_b, threshold)
        sens_list.append(s)
        spec_list.append(sp)
        ppv_list.append(p)
        npv_list.append(n_)

    def _ci(arr):
        return np.percentile(arr, [2.5, 97.5])

    sens_lo, sens_hi = _ci(sens_list)
    spec_lo, spec_hi = _ci(spec_list)
    ppv_lo, ppv_hi = _ci(ppv_list)
    npv_lo, npv_hi = _ci(npv_list)
    return (sens_lo, sens_hi), (spec_lo, spec_hi), (ppv_lo, ppv_hi), (npv_lo, npv_hi)


# ==========================================
# 训练辅助函数
# ==========================================

def train_cls_stream(X_train, y_train, X_val, y_val, X_test, use_gpu=True):
    """Stream 1: 分类流 XGBoost"""
    device_str = 'cuda' if use_gpu and torch.cuda.is_available() else 'cpu'
    dtrain = xgb.DMatrix(X_train, label=y_train)
    dval = xgb.DMatrix(X_val, label=y_val)
    dtest = xgb.DMatrix(X_test)
    params = {
        'max_depth': 4, 'learning_rate': 0.03, 'subsample': 0.7,
        'colsample_bytree': 0.7, 'tree_method': 'hist',
        'device': device_str, 'disable_default_eval_metric': 1, 'eval_metric': 'auc',
    }
    model = xgb.train(params, dtrain, num_boost_round=1200, obj=asymmetric_focal_loss_obj,
                       evals=[(dval, 'val')], early_stopping_rounds=100, verbose_eval=False)
    pred_val = expit(model.predict(dval))
    pred_test = expit(model.predict(dtest))
    return model, pred_val, pred_test


def train_rank_stream(X_train, y_train, X_val, X_test, use_gpu=True):
    """Stream 2: 排序流 XGBoost"""
    device_str = 'cuda' if use_gpu and torch.cuda.is_available() else 'cpu'
    X_rank, y_rank, groups = make_ranking_data(X_train, y_train)
    dtrain = xgb.DMatrix(X_rank, label=y_rank)
    dtrain.set_group(groups)
    dval = xgb.DMatrix(X_val)
    dtest = xgb.DMatrix(X_test)
    params = {
        'objective': 'rank:pairwise', 'eval_metric': 'auc',
        'max_depth': 5, 'learning_rate': 0.05, 'subsample': 0.8,
        'tree_method': 'hist', 'device': device_str,
    }
    model = xgb.train(params, dtrain, num_boost_round=800, verbose_eval=False)
    return model, model.predict(dval), model.predict(dtest)


def teacher_fusion(pred_cls_val, pred_cls_test, pred_rnk_val, pred_rnk_test, y_val):
    """Teacher 融合 (rank norm + linear fusion on val)"""
    rp_cls_val = rankdata(pred_cls_val) / len(pred_cls_val)
    rp_rnk_val = rankdata(pred_rnk_val) / len(pred_rnk_val)
    rp_cls_test = np.array([percentileofscore(pred_cls_val, x, kind='weak') / 100 for x in pred_cls_test])
    rp_rnk_test = np.array([percentileofscore(pred_rnk_val, x, kind='weak') / 100 for x in pred_rnk_test])
    w, _ = find_best_fusion_weight(rp_cls_val, rp_rnk_val, y_val)
    teacher_val = w * rp_cls_val + (1 - w) * rp_rnk_val
    teacher_test = w * rp_cls_test + (1 - w) * rp_rnk_test
    return teacher_val, teacher_test, w


def train_student(X_train, y_train, X_val, y_val, X_test,
                  soft_labels_train, device, epochs=150):
    """Student MLP 知识蒸馏"""
    n_feat = X_train.shape[1]
    X_tr_t = torch.FloatTensor(X_train.values).to(device)
    y_tr_t = torch.FloatTensor(y_train.values).to(device)
    t_soft = torch.FloatTensor(soft_labels_train).reshape(-1, 1).to(device)
    X_va_t = torch.FloatTensor(X_val.values).to(device)
    X_te_t = torch.FloatTensor(X_test.values).to(device)

    ds = TensorDataset(X_tr_t, y_tr_t, t_soft)
    dl = DataLoader(ds, batch_size=1024, shuffle=True)

    student = StudentNet(n_feat).to(device)
    optimizer = optim.AdamW(student.parameters(), lr=0.001, weight_decay=1e-5)
    scheduler = optim.lr_scheduler.CosineAnnealingLR(optimizer, T_max=epochs)
    bce = nn.BCEWithLogitsLoss()
    mse = nn.MSELoss()

    best_auc, best_state = 0, None
    for epoch in range(epochs):
        student.train()
        for bx, by, bt in dl:
            optimizer.zero_grad()
            logits = student(bx)
            probs = torch.sigmoid(logits)
            loss = 0.2 * bce(logits, by.reshape(-1, 1)) + 0.8 * mse(probs, bt)
            loss.backward()
            optimizer.step()
        scheduler.step()
        if (epoch + 1) % 10 == 0:
            student.eval()
            with torch.no_grad():
                va_p = torch.sigmoid(student(X_va_t)).cpu().numpy().flatten()
                auc = roc_auc_score(y_val, va_p)
                if auc > best_auc:
                    best_auc = auc
                    best_state = copy.deepcopy(student.state_dict())

    if best_state:
        student.load_state_dict(best_state)
    student.eval()
    with torch.no_grad():
        pred_val = torch.sigmoid(student(X_va_t)).cpu().numpy().flatten()
        pred_test = torch.sigmoid(student(X_te_t)).cpu().numpy().flatten()
    return pred_val, pred_test, best_auc


def train_mlp_direct(X_train, y_train, X_val, y_val, X_test, device, epochs=150):
    """直接 MLP (无蒸馏，纯硬标签训练) — 用于消融对比"""
    n_feat = X_train.shape[1]
    X_tr_t = torch.FloatTensor(X_train.values).to(device)
    y_tr_t = torch.FloatTensor(y_train.values).to(device)
    X_va_t = torch.FloatTensor(X_val.values).to(device)
    X_te_t = torch.FloatTensor(X_test.values).to(device)

    ds = TensorDataset(X_tr_t, y_tr_t)
    dl = DataLoader(ds, batch_size=1024, shuffle=True)

    model = StudentNet(n_feat).to(device)
    # 使用加权 BCE 处理不平衡
    pos_w = float((y_train == 0).sum()) / float((y_train == 1).sum())
    bce = nn.BCEWithLogitsLoss(pos_weight=torch.tensor([pos_w]).to(device))
    optimizer = optim.AdamW(model.parameters(), lr=0.001, weight_decay=1e-5)
    scheduler = optim.lr_scheduler.CosineAnnealingLR(optimizer, T_max=epochs)

    best_auc, best_state = 0, None
    for epoch in range(epochs):
        model.train()
        for bx, by in dl:
            optimizer.zero_grad()
            logits = model(bx)
            loss = bce(logits, by.reshape(-1, 1))
            loss.backward()
            optimizer.step()
        scheduler.step()
        if (epoch + 1) % 10 == 0:
            model.eval()
            with torch.no_grad():
                va_p = torch.sigmoid(model(X_va_t)).cpu().numpy().flatten()
                auc = roc_auc_score(y_val, va_p)
                if auc > best_auc:
                    best_auc = auc
                    best_state = copy.deepcopy(model.state_dict())

    if best_state:
        model.load_state_dict(best_state)
    model.eval()
    with torch.no_grad():
        pred_val = torch.sigmoid(model(X_va_t)).cpu().numpy().flatten()
        pred_test = torch.sigmoid(model(X_te_t)).cpu().numpy().flatten()
    return pred_val, pred_test, best_auc


# ==========================================
# 评估辅助
# ==========================================
def evaluate_ablation(name, y_val, p_val, y_test, p_test):
    th = youden_threshold(y_val, p_val)
    auc = roc_auc_score(y_test, p_test)
    auc_lo, auc_hi = bootstrap_auc_ci(y_test, p_test)
    sens, spec, ppv, npv = clinical_metrics(y_test, p_test, th)
    (sens_lo, sens_hi), (spec_lo, spec_hi), (ppv_lo, ppv_hi), (npv_lo, npv_hi) = bootstrap_metric_ci(
        y_test, p_test, threshold=th
    )

    def fmt(x, lo, hi):
        return f"{x:.4f} ({lo:.4f}–{hi:.4f})"

    return {
        "消融配置": name,
        "Threshold(Val Youden)": round(float(th), 4),
        "AUC (95% CI)": fmt(auc, auc_lo, auc_hi),
        "Sensitivity (95% CI)": fmt(sens, sens_lo, sens_hi),
        "Specificity (95% CI)": fmt(spec, spec_lo, spec_hi),
        "PPV (95% CI)": fmt(ppv, ppv_lo, ppv_hi),
        "NPV (95% CI)": fmt(npv, npv_lo, npv_hi),
    }


# ==========================================
# 主流程
# ==========================================
def main():
    np.random.seed(config.SEED)
    torch.manual_seed(config.SEED)
    use_gpu = torch.cuda.is_available()
    device = torch.device("cuda" if use_gpu else "cpu")

    print("\n" + "=" * 70)
    print("   [Step 6] Ablation Study — 消融实验")
    print("=" * 70)

    # 加载数据
    train = pd.read_csv(config.PROCESSED_DATA_PREFIX + "train.csv")
    val = pd.read_csv(config.PROCESSED_DATA_PREFIX + "val.csv")
    test = pd.read_csv(config.PROCESSED_DATA_PREFIX + "test.csv")

    all_features = [c for c in train.columns if c not in [config.TARGET_COL, 'cohortid']]
    raw_features = [c for c in all_features if not c.startswith("Embedding_")]
    embed_features = [c for c in all_features if c.startswith("Embedding_")]

    y_train, y_val, y_test = train[config.TARGET_COL], val[config.TARGET_COL], test[config.TARGET_COL]

    print(f"   全部特征: {len(all_features)}  (原始: {len(raw_features)}, 嵌入: {len(embed_features)})")
    print(f"   Device: {device}\n")

    results = []

    # ========================================
    # A1. 仅原始特征 + 单流分类 (最简基线)
    # ========================================
    print("   [A1] 仅原始特征  + 单流分类 XGBoost ...")
    X_tr_raw, X_va_raw, X_te_raw = train[raw_features], val[raw_features], test[raw_features]
    _, cls_val_raw, cls_test_raw = train_cls_stream(X_tr_raw, y_train, X_va_raw, y_val, X_te_raw, use_gpu)
    auc_a1 = roc_auc_score(y_test, cls_test_raw)
    results.append(evaluate_ablation("A1: 原始特征 + 单流XGB", y_val, cls_val_raw, y_test, cls_test_raw))
    print(f"        AUC = {auc_a1:.4f}")

    # ========================================
    # A2. 全部特征 + 单流分类 XGBoost
    # ========================================
    print("   [A2] 全部特征  + 单流分类 XGBoost ...")
    X_tr_all, X_va_all, X_te_all = train[all_features], val[all_features], test[all_features]
    model_cls_all, cls_val_all, cls_test_all = train_cls_stream(X_tr_all, y_train, X_va_all, y_val, X_te_all, use_gpu)
    auc_a2 = roc_auc_score(y_test, cls_test_all)
    results.append(evaluate_ablation("A2: 全部特征 + 单流XGB", y_val, cls_val_all, y_test, cls_test_all))
    print(f"        AUC = {auc_a2:.4f}")

    # ========================================
    # A3. 全部特征 + 直接 MLP
    # ========================================
    print("   [A3] 全部特征  + 直接 MLP ...")
    mlp_val, mlp_test, mlp_best_auc = train_mlp_direct(
        X_tr_all, y_train, X_va_all, y_val, X_te_all, device
    )
    auc_a3 = roc_auc_score(y_test, mlp_test)
    results.append(evaluate_ablation("A3: 全部特征 + 直接MLP", y_val, mlp_val, y_test, mlp_test))
    print(f"        AUC = {auc_a3:.4f}  (val best = {mlp_best_auc:.4f})")

    # ========================================
    # A4/A5. 复用 0313 已保存结果（Teacher/Student）
    # ========================================
    reuse_dir = os.path.join(config.BASE_DIR, "results_flat_0313")
    val_pred_path = os.path.join(reuse_dir, "val_predictions.csv")
    test_pred_path = os.path.join(reuse_dir, "final_predictions.csv")

    print("   [A4/A5] 复用 0313 的 Teacher/Student 预测结果 ...")
    if not (os.path.exists(val_pred_path) and os.path.exists(test_pred_path)):
        raise FileNotFoundError(f"未找到 0313 预测文件: {val_pred_path} 或 {test_pred_path}")

    val_pred = pd.read_csv(val_pred_path)
    test_pred = pd.read_csv(test_pred_path)

    need_cols = {"target", "pred_teacher", "pred_student"}
    if not need_cols.issubset(val_pred.columns) or not need_cols.issubset(test_pred.columns):
        raise ValueError("0313 预测文件缺少必需列: target, pred_teacher, pred_student")

    y_val_ref = val_pred["target"].values
    y_test_ref = test_pred["target"].values

    teacher_val = val_pred["pred_teacher"].values
    teacher_test = test_pred["pred_teacher"].values
    auc_a4 = roc_auc_score(y_test_ref, teacher_test)
    results.append(evaluate_ablation("A4: 全部特征 + Teacher", y_val_ref, teacher_val, y_test_ref, teacher_test))
    print(f"        Teacher AUC = {auc_a4:.4f}")

    student_val = val_pred["pred_student"].values
    student_test = test_pred["pred_student"].values
    auc_a5 = roc_auc_score(y_test_ref, student_test)
    results.append(evaluate_ablation("A5: 全部特征 + Student", y_val_ref, student_val, y_test_ref, student_test))
    print(f"        Student AUC = {auc_a5:.4f}")

    # ========================================
    # 汇总
    # ========================================
    print("\n" + "=" * 100)
    print("   消融实验结果汇总")
    print("=" * 100)
    print(f"   {'消融配置':<30} {'AUC (95% CI)':<26} {'Sensitivity (95% CI)':<26} {'Specificity (95% CI)':<26}")
    print(f"   {'':<30} {'PPV (95% CI)':<26} {'NPV (95% CI)':<26} {'Threshold':<12}")
    print("   " + "-" * 100)
    for r in results:
        print(f"   {r['消融配置']:<30} {r['AUC (95% CI)']:<26} {r['Sensitivity (95% CI)']:<26} {r['Specificity (95% CI)']:<26}")
        print(f"   {'':<30} {r['PPV (95% CI)']:<26} {r['NPV (95% CI)']:<26} {r['Threshold(Val Youden)']:<12.4f}")
    print("   " + "-" * 100)

    # 保存
    os.makedirs(config.OUTPUT_DIR, exist_ok=True)
    csv_path = os.path.join(config.OUTPUT_DIR, "ablation_table.csv")
    pd.DataFrame(results).to_csv(csv_path, index=False, encoding="utf-8-sig")
    print(f"\n   [SAVED] {csv_path}")

    # Excel 格式
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        xl_path = os.path.join(config.OUTPUT_DIR, "ablation_table.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ablation"

        headers = [
            "消融配置",
            "Threshold(Val Youden)",
            "AUC (95% CI)",
            "Sensitivity (95% CI)",
            "Specificity (95% CI)",
            "PPV (95% CI)",
            "NPV (95% CI)",
        ]
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin = Border(*(Side(style="thin"),) * 4)

        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        student_fill = PatternFill("solid", fgColor="E2EFDA")
        for ri, r in enumerate(results, 2):
            is_student = "Student" in r["消融配置"]
            for ci, key in enumerate(headers, 1):
                val = r[key]
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.alignment = Alignment(horizontal="center")
                if is_student:
                    cell.fill = student_fill
                    cell.font = Font(bold=True)

        for i, w in enumerate([30, 20, 26, 26, 26, 20, 20], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        wb.save(xl_path)
        print(f"   [SAVED] {xl_path}")
    except Exception as e:
        print(f"   [WARN] Excel 保存失败: {e}")

    print("\n   ✅ 消融实验完成！\n")


if __name__ == "__main__":
    main()
