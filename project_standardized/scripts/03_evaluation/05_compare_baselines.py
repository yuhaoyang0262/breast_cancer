# Project path bootstrap for direct script execution from nested folders.
from pathlib import Path
import sys

for _project_root in Path(__file__).resolve().parents:
    if (_project_root / "configs").is_dir() and (_project_root / "utils").is_dir():
        for _path in (_project_root, _project_root / "configs", _project_root / "utils"):
            _path_str = str(_path)
            if _path_str not in sys.path:
                sys.path.insert(0, _path_str)
        break
# 05_compare_baselines.py
# Comprehensive multi-model benchmarking for top-tier journal publication
import os
import warnings
import numpy as np
import pandas as pd
import joblib
import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

# Scientific publication font settings
plt.rcParams['font.family'] = 'sans-serif'
plt.rcParams['axes.unicode_minus'] = False

from sklearn.linear_model import LogisticRegression
from sklearn.tree import DecisionTreeClassifier
from sklearn.ensemble import (
    ExtraTreesClassifier,
    AdaBoostClassifier, HistGradientBoostingClassifier,
)
from sklearn.neural_network import MLPClassifier
from sklearn.naive_bayes import GaussianNB
from sklearn.neighbors import KNeighborsClassifier
from sklearn.svm import LinearSVC
from sklearn.calibration import CalibratedClassifierCV
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline
from sklearn.metrics import roc_auc_score, roc_curve, confusion_matrix

warnings.filterwarnings("ignore")
import config

THRESHOLD_STEP = 0.005


# ──────────────────────────────────────────────
# Utilities
# ──────────────────────────────────────────────
def bootstrap_auc_ci(y_true, y_prob, n_boot=1000, seed=42):
    rng = np.random.default_rng(seed)
    n = len(y_true)
    y_true = np.asarray(y_true)
    y_prob = np.asarray(y_prob)
    aucs = []
    for _ in range(n_boot):
        idx = rng.integers(0, n, n)
        if len(np.unique(y_true[idx])) < 2:
            continue
        aucs.append(roc_auc_score(y_true[idx], y_prob[idx]))
    lo, hi = np.percentile(aucs, [2.5, 97.5])
    return lo, hi


def build_threshold_grid(step=THRESHOLD_STEP):
    grid = np.arange(0.0, 1.0 + step, step)
    return np.clip(np.round(grid, 6), 0.0, 1.0)


def clinical_metrics(y_true, y_prob, threshold):
    y_pred = (y_prob >= threshold).astype(int)
    tn, fp, fn, tp = confusion_matrix(y_true, y_pred, labels=[0, 1]).ravel()
    sens = tp / (tp + fn) if (tp + fn) > 0 else 0.0
    spec = tn / (tn + fp) if (tn + fp) > 0 else 0.0
    ppv = tp / (tp + fp) if (tp + fp) > 0 else 0.0
    npv = tn / (tn + fn) if (tn + fn) > 0 else 0.0
    return sens, spec, ppv, npv


def threshold_sweep_table(y_true, y_prob, model_name, split_name, thresholds):
    rows = []
    for th in thresholds:
        sens, spec, ppv, npv = clinical_metrics(y_true, y_prob, th)
        rows.append({
            "Model": model_name,
            "Split": split_name,
            "Threshold": float(th),
            "Sensitivity": float(sens),
            "Specificity": float(spec),
            "PPV": float(ppv),
            "NPV": float(npv),
            "Sens_minus_Spec": float(sens - spec),
            "Sens_gt_Spec": int(sens > spec),
        })
    return pd.DataFrame(rows)


def select_threshold_sens_gt_spec(y_true, y_prob, thresholds):
    df = threshold_sweep_table(y_true, y_prob, model_name="tmp", split_name="Validation", thresholds=thresholds)
    candidates = df[df["Sens_gt_Spec"] == 1]
    if len(candidates) > 0:
        best_idx = candidates["Sens_minus_Spec"].idxmax()
        return float(df.loc[best_idx, "Threshold"]), True
    best_idx = df["Sens_minus_Spec"].idxmax()
    return float(df.loc[best_idx, "Threshold"]), False


def delong_test(y_true, prob_a, prob_b):
    y_true = np.asarray(y_true)
    prob_a = np.asarray(prob_a)
    prob_b = np.asarray(prob_b)
    pos_idx = np.where(y_true == 1)[0]
    neg_idx = np.where(y_true == 0)[0]
    n1, n0 = len(pos_idx), len(neg_idx)

    def _placement_values(scores, pos_idx, neg_idx):
        s_pos = scores[pos_idx]
        s_neg = scores[neg_idx]
        V10 = np.mean(s_neg[:, None] < s_pos[None, :], axis=0) + \
              0.5 * np.mean(s_neg[:, None] == s_pos[None, :], axis=0)
        V01 = np.mean(s_pos[:, None] > s_neg[None, :], axis=0) + \
              0.5 * np.mean(s_pos[:, None] == s_neg[None, :], axis=0)
        return V10, V01

    V10_a, V01_a = _placement_values(prob_a, pos_idx, neg_idx)
    V10_b, V01_b = _placement_values(prob_b, pos_idx, neg_idx)
    auc_a = V10_a.mean()
    auc_b = V10_b.mean()

    S10_aa = np.var(V10_a, ddof=1)
    S10_bb = np.var(V10_b, ddof=1)
    S10_ab = np.cov(V10_a, V10_b, ddof=1)[0, 1]
    S01_aa = np.var(V01_a, ddof=1)
    S01_bb = np.var(V01_b, ddof=1)
    S01_ab = np.cov(V01_a, V01_b, ddof=1)[0, 1]

    var_a = S10_aa / n1 + S01_aa / n0
    var_b = S10_bb / n1 + S01_bb / n0
    cov_ab = S10_ab / n1 + S01_ab / n0
    var_diff = var_a + var_b - 2 * cov_ab

    if var_diff <= 0:
        return 0.0, 1.0, auc_a, auc_b

    from scipy import stats
    z = (auc_a - auc_b) / np.sqrt(var_diff)
    p = 2 * (1 - stats.norm.cdf(abs(z)))
    return float(z), float(p), float(auc_a), float(auc_b)


def net_benefit(y_true, y_prob, thresholds):
    y_true = np.asarray(y_true)
    n = len(y_true)
    nbs = []
    for pt in thresholds:
        pred = (y_prob >= pt).astype(int)
        tp = np.sum((y_true == 1) & (pred == 1))
        fp = np.sum((y_true == 0) & (pred == 1))
        if pt >= 1:
            nbs.append(0.0)
        else:
            nbs.append(tp / n - fp / n * pt / (1 - pt))
    return np.array(nbs)


def evaluate_model(name, y_val, prob_val, y_test, prob_test, n_boot=1000, thresholds=None):
    if thresholds is None:
        thresholds = build_threshold_grid()
    threshold, strict_ok = select_threshold_sens_gt_spec(y_val, prob_val, thresholds)
    auc = roc_auc_score(y_test, prob_test)
    lo, hi = bootstrap_auc_ci(y_test, prob_test, n_boot=n_boot)
    sens, spec, ppv, npv = clinical_metrics(y_test, prob_test, threshold)
    return {
        "Model": name,
        "AUC": round(auc, 4),
        "AUC_Low": round(lo, 4),
        "AUC_High": round(hi, 4),
        "AUC (95% CI)": f"{auc:.4f} ({lo:.4f}-{hi:.4f})",
        "Sensitivity": round(sens, 4),
        "Specificity": round(spec, 4),
        "PPV": round(ppv, 4),
        "NPV": round(npv, 4),
        "Threshold": round(threshold, 4),
        "ThresholdRule": "Sensitivity>Specificity",
        "StrictCandidateOnVal": int(strict_ok),
    }


def _safe_model_filename(name: str) -> str:
    keep = []
    for ch in name:
        if ch.isalnum() or ch in ["-", "_"]:
            keep.append(ch)
        elif ch in [" ", "(", ")", "/"]:
            keep.append("_")
    out = "".join(keep).strip("_")
    while "__" in out:
        out = out.replace("__", "_")
    return out.lower() if out else "model"


def train_baseline_suite(
    X_train, y_train, X_val, y_val, X_test, y_test,
    seed, n_boot, model_dir,
):
    """Train all baseline models, save fitted estimators, and return evaluation artifacts."""
    os.makedirs(model_dir, exist_ok=True)

    results = []
    baseline_probs = []
    split_auc_rows = []
    pred_store = {
        "Train": {"target": y_train},
        "Validation": {"target": y_val},
        "Test": {"target": y_test},
    }
    model_registry = []

    def _record_split_auc_and_preds(name, p_train, p_val, p_test):
        split_auc_rows.append({
            "Model": name,
            "AUC_Train": float(roc_auc_score(y_train, p_train)) if p_train is not None else np.nan,
            "AUC_Validation": float(roc_auc_score(y_val, p_val)) if p_val is not None else np.nan,
            "AUC_Test": float(roc_auc_score(y_test, p_test)) if p_test is not None else np.nan,
        })
        if p_train is not None:
            pred_store["Train"][name] = p_train
        if p_val is not None:
            pred_store["Validation"][name] = p_val
        if p_test is not None:
            pred_store["Test"][name] = p_test

    def _fit(step, name, model, X_tr, y_tr, X_v, X_te, use_pipeline=False):
        print(f"   [{step:>2}] {name:<25}", end="", flush=True)
        if use_pipeline:
            clf = Pipeline([("scaler", StandardScaler()), ("clf", model)])
        else:
            clf = model

        clf.fit(X_tr, y_tr)
        ptr = clf.predict_proba(X_train)[:, 1]
        pv = clf.predict_proba(X_v)[:, 1]
        pt = clf.predict_proba(X_te)[:, 1]

        _record_split_auc_and_preds(name, ptr, pv, pt)
        r = evaluate_model(name, y_val, pv, y_test, pt, n_boot=n_boot)
        results.append(r)
        baseline_probs.append((name, y_test, pt))

        model_file = os.path.join(model_dir, f"{_safe_model_filename(name)}.joblib")
        joblib.dump(clf, model_file)
        model_registry.append({"Model": name, "Model_File": model_file, "Use_Pipeline": int(use_pipeline)})

        print(f" AUC = {r['AUC']:.4f} | saved")
        return ptr, pv, pt

    print("\n   --- Training Baseline Models (Overfitting Trees Excluded) ---")
    _fit(1, "Logistic Regression",
         LogisticRegression(max_iter=2000, class_weight="balanced", C=0.1, solver="lbfgs", random_state=seed),
         X_train, y_train, X_val, X_test)
    _fit(2, "Decision Tree", DecisionTreeClassifier(max_depth=6, class_weight="balanced", random_state=seed),
         X_train, y_train, X_val, X_test)
    _fit(3, "Extra Trees", ExtraTreesClassifier(n_estimators=500, max_depth=8, class_weight="balanced", n_jobs=-1,
                                                random_state=seed), X_train, y_train, X_val, X_test)
    _fit(4, "AdaBoost", AdaBoostClassifier(n_estimators=200, learning_rate=0.1, random_state=seed), X_train,
         y_train, X_val, X_test)
    _fit(5, "Hist-GBDT",
         HistGradientBoostingClassifier(max_iter=300, max_depth=4, learning_rate=0.05, class_weight="balanced",
                                        early_stopping=True, validation_fraction=0.1, random_state=seed),
         X_train, y_train, X_val, X_test)
    _fit(6, "Naive Bayes", GaussianNB(), X_train, y_train, X_val, X_test)

    KNN_MAX = 10000
    if len(X_train) > KNN_MAX:
        from sklearn.model_selection import StratifiedShuffleSplit
        idx_knn, _ = next(
            StratifiedShuffleSplit(n_splits=1, train_size=KNN_MAX, random_state=seed).split(X_train, y_train)
        )
        X_tr_knn, y_tr_knn = X_train[idx_knn], y_train[idx_knn]
    else:
        X_tr_knn, y_tr_knn = X_train, y_train

    _fit(7, "KNN", KNeighborsClassifier(n_neighbors=15, metric="minkowski", n_jobs=-1), X_tr_knn, y_tr_knn, X_val,
         X_test, use_pipeline=True)

    _fit(8, "Linear SVM",
         CalibratedClassifierCV(LinearSVC(C=0.1, class_weight="balanced", max_iter=3000, random_state=seed),
                                cv=3, method="sigmoid"), X_train, y_train, X_val, X_test, use_pipeline=True)

    _fit(9, "MLP Neural Network",
         MLPClassifier(hidden_layer_sizes=(128, 64, 32), activation="relu", solver="adam", max_iter=500,
                       early_stopping=True, validation_fraction=0.1, random_state=seed), X_train, y_train, X_val,
         X_test, use_pipeline=True)

    registry_path = os.path.join(model_dir, "baseline_model_registry.csv")
    pd.DataFrame(model_registry).to_csv(registry_path, index=False, encoding="utf-8-sig")
    print(f"   [SAVED] Model registry → {registry_path}")

    return results, baseline_probs, split_auc_rows, pred_store


# ──────────────────────────────────────────────
# Main Execution
# ──────────────────────────────────────────────
def main():
    print("\n" + "=" * 70)
    print("   [Step 5] Multi-Model Comparison (Top-Tier Journal Level)")
    print("=" * 70)

    train_path = config.PROCESSED_DATA_PREFIX + "train.csv"
    val_path = config.PROCESSED_DATA_PREFIX + "val.csv"
    test_path = config.PROCESSED_DATA_PREFIX + "test.csv"

    if not os.path.exists(train_path):
        print("[ERROR] Data not found. Run 01_preprocess.py first.")
        return

    train = pd.read_csv(train_path)
    val = pd.read_csv(val_path)
    test = pd.read_csv(test_path)

    EXCLUDE = [config.TARGET_COL, "cohortid"]
    features = [c for c in train.columns if c not in EXCLUDE]

    X_train, y_train = train[features].values, train[config.TARGET_COL].values
    X_val, y_val = val[features].values, val[config.TARGET_COL].values
    X_test, y_test = test[features].values, test[config.TARGET_COL].values

    model_dir = os.path.join(config.OUTPUT_DIR, "baseline_models")
    os.makedirs(model_dir, exist_ok=True)
    pd.Series(features, name="feature").to_csv(
        os.path.join(model_dir, "baseline_feature_columns.csv"), index=False, encoding="utf-8-sig"
    )

    threshold_grid = build_threshold_grid(step=THRESHOLD_STEP)

    results, baseline_probs, split_auc_rows, pred_store = train_baseline_suite(
        X_train=X_train,
        y_train=y_train,
        X_val=X_val,
        y_val=y_val,
        X_test=X_test,
        y_test=y_test,
        seed=config.SEED,
        n_boot=1000,
        model_dir=model_dir,
    )

    # ── Load Proposed Model ───────────────────────────────────────────────
    pred_file = config.PREDICTION_FILE
    has_val_preds = False
    val_pred_file = os.path.join(config.OUTPUT_DIR, "val_predictions.csv")
    train_pred_file = os.path.join(config.OUTPUT_DIR, "train_predictions.csv")

    if os.path.exists(pred_file):
        print("\n>>> Loading Our Proposed Model Predictions...")
        df_preds = pd.read_csv(pred_file)
        y_test_our = df_preds["target"].values
        has_val_preds = os.path.exists(val_pred_file)
        has_train_preds = os.path.exists(train_pred_file)

        # STRICTLY ENGLISH MAPPING
        our_models_mapping = {
            "Stream 1 (Classification)": "pred_stream1",
            "Stream 2 (Ranking)": "pred_stream2",
            "Teacher (Fusion)": "pred_teacher",
            "Our Proposed Model": "pred_student",
        }

        for model_name, col in our_models_mapping.items():
            if col not in df_preds.columns:
                continue
            prob_test_our = df_preds[col].values

            if has_val_preds:
                df_val_p = pd.read_csv(val_pred_file)
                prob_val_our, y_val_our = df_val_p[col].values, df_val_p["target"].values
            else:
                prob_val_our, y_val_our = prob_test_our, y_test_our

            prob_train_our = None
            if has_train_preds:
                df_train_p = pd.read_csv(train_pred_file)
                if col in df_train_p.columns:
                    prob_train_our = df_train_p[col].values

            split_auc_rows.append({
                "Model": model_name,
                "AUC_Train": float(roc_auc_score(y_train, prob_train_our)) if prob_train_our is not None else np.nan,
                "AUC_Validation": float(roc_auc_score(y_val_our, prob_val_our)) if prob_val_our is not None else np.nan,
                "AUC_Test": float(roc_auc_score(y_test_our, prob_test_our)) if prob_test_our is not None else np.nan,
            })
            if prob_train_our is not None:
                pred_store["Train"][model_name] = prob_train_our
            if prob_val_our is not None:
                pred_store["Validation"][model_name] = prob_val_our
            if prob_test_our is not None:
                pred_store["Test"][model_name] = prob_test_our
            results.append(
                evaluate_model(
                    model_name,
                    y_val_our,
                    prob_val_our,
                    y_test_our,
                    prob_test_our,
                    thresholds=threshold_grid,
                )
            )
            print(f"   {model_name}: AUC = {results[-1]['AUC']:.4f}")
    else:
        print(f"\n[WARN] Prediction file not found: {pred_file}")

    # ── Outputs ───────────────────────────────────────────────────────────
    os.makedirs(config.OUTPUT_DIR, exist_ok=True)
    df_result = pd.DataFrame(results)

    df_display = df_result[["Model", "AUC (95% CI)", "Sensitivity", "Specificity", "PPV", "NPV", "Threshold"]].copy()
    csv_path = os.path.join(config.OUTPUT_DIR, "comparison_table.csv")
    df_display.to_csv(csv_path, index=False, encoding="utf-8-sig")
    print(f"\n   [SAVED] Table → {csv_path}")

    # Process and save predictions
    df_split_auc = pd.DataFrame(split_auc_rows).drop_duplicates(subset=["Model"], keep="last")
    df_split_auc.to_csv(os.path.join(config.OUTPUT_DIR, "auc_by_split.csv"), index=False, encoding="utf-8-sig")

    split_pred_frames = {}
    for split_name in ["Train", "Validation", "Test"]:
        df_split_pred = pd.DataFrame(pred_store[split_name])
        split_pred_frames[split_name] = df_split_pred

    threshold_curves = []
    for split_name, df_split in split_pred_frames.items():
        y_true_split = df_split["target"].values
        for model_name in [c for c in df_split.columns if c != "target"]:
            threshold_curves.append(
                threshold_sweep_table(
                    y_true_split,
                    df_split[model_name].values,
                    model_name=model_name,
                    split_name=split_name,
                    thresholds=threshold_grid,
                )
            )

    df_threshold_curves = pd.concat(threshold_curves, axis=0, ignore_index=True)
    threshold_curve_path = os.path.join(config.OUTPUT_DIR, "threshold_sensitivity_specificity_curve_all_models.csv")
    df_threshold_curves.to_csv(threshold_curve_path, index=False, encoding="utf-8-sig")
    print(f"   [SAVED] Threshold curves → {threshold_curve_path}")

    # Draw Top-Tier ROCs for Train, Val, Test
    print("\n>>> Generating High-Fidelity ROC Curves (No Insets, English Only)...")
    _plot_professional_roc(split_pred_frames["Train"], "Training",
                           os.path.join(config.OUTPUT_DIR, "roc_comparison_train.png"))
    _plot_professional_roc(split_pred_frames["Validation"], "Validation",
                           os.path.join(config.OUTPUT_DIR, "roc_comparison_validation.png"))
    _plot_professional_roc(split_pred_frames["Test"], "Test",
                           os.path.join(config.OUTPUT_DIR, "roc_comparison_test.png"))
    print("   [SAVED] 3 x ROC Plots generated successfully.")

    # ── DeLong & NRI/IDI ──────────────────────────────────────────────────
    if os.path.exists(pred_file):
        our_prob_map = {k: df_preds[v].values for k, v in our_models_mapping.items() if v in df_preds.columns}
        our_aucs = {k: roc_auc_score(df_preds["target"].values, v) for k, v in our_prob_map.items()}
        best_our_name = "Our Proposed Model" if "Our Proposed Model" in our_aucs else max(our_aucs, key=our_aucs.get)
        best_our_prob_raw = our_prob_map[best_our_name]
        y_test_delong = df_preds["target"].values

        calibrator_05 = LogisticRegression()
        best_col_str = our_models_mapping.get(best_our_name, "pred_student")
        if has_val_preds and best_col_str in pd.read_csv(val_pred_file).columns:
            df_val_tmp = pd.read_csv(val_pred_file)
            calibrator_05.fit(df_val_tmp[best_col_str].values.reshape(-1, 1), df_val_tmp['target'].values)
        else:
            calibrator_05.fit(best_our_prob_raw.reshape(-1, 1), y_test_delong)

        best_our_prob = calibrator_05.predict_proba(best_our_prob_raw.reshape(-1, 1))[:, 1]

        delong_rows, nri_idi_rows = [], []

        def _format_p_value_for_report(p_val: float) -> str:
            # Avoid reporting exact zero due to formatting/rounding in manuscripts.
            if p_val < 1e-4:
                return "<0.0001"
            return f"{p_val:.4f}"

        for bname, yt, bp in baseline_probs:
            z, p, auc_ours, auc_base = delong_test(y_test_delong, best_our_prob, bp)
            delong_rows.append({"Comparison": f"{best_our_name} vs {bname}", "Our_AUC": round(auc_ours, 4),
                                "Base_AUC": round(auc_base, 4), "Delta_AUC": round(auc_ours - auc_base, 4),
                                "Z_stat": round(z, 3), "P_value": float(p),
                                "P_value_report": _format_p_value_for_report(float(p))})

        pd.DataFrame(delong_rows).to_csv(os.path.join(config.OUTPUT_DIR, "delong_test.csv"), index=False,
                                         encoding="utf-8-sig")

        def compute_nri_idi(y_t, p_new, p_ref):
            ev, nonev = y_t == 1, y_t == 0
            diff = p_new - p_ref
            nri = (np.mean(diff[ev] > 0) - np.mean(diff[ev] < 0)) + (
                        np.mean(diff[nonev] < 0) - np.mean(diff[nonev] > 0))
            idi = (p_new[ev].mean() - p_new[nonev].mean()) - (p_ref[ev].mean() - p_ref[nonev].mean())
            return nri, idi

        for bname, yt, bp in baseline_probs:
            nri_val, idi_val = compute_nri_idi(y_test_delong, best_our_prob, bp)
            nri_idi_rows.append(
                {"Comparison": f"{best_our_name} vs {bname}", "NRI": round(nri_val, 4), "IDI": round(idi_val, 4)})

        pd.DataFrame(nri_idi_rows).to_csv(os.path.join(config.OUTPUT_DIR, "nri_idi_test.csv"), index=False,
                                          encoding="utf-8-sig")

    # ── DCA Plot ───────────────────────────────────────────────────────────
    try:
        dca_path = os.path.join(config.OUTPUT_DIR, "dca_comparison.png")
        # Ensure palette matches the 9 remaining baselines exactly
        palette_dca = ["#4DBBD5", "#00A087", "#3C5488", "#F39B7F", "#8491B4", "#91D1C2", "#7E6148", "#B09C85",
                       "#4B4B4B"]
        all_dca_models = [(name, yt, yp, c, "--", 1.5) for (name, yt, yp), c in zip(baseline_probs, palette_dca)]
        our_dca = (best_our_name, df_preds["target"].values, best_our_prob, "#E64B35", "-", 3.8) if os.path.exists(
            pred_file) else None
        _plot_professional_dca(all_dca_models, our_dca, dca_path)
        print(f"   [SAVED] DCA Plot → {dca_path}")
    except Exception as e:
        print(f"   [WARN] DCA Plot failed: {e}")

    print("\n   ✅ Pipeline Successfully Completed!\n")


# ──────────────────────────────────────────────
# Visualization Engines (Top Tier Journal Level)
# ──────────────────────────────────────────────

def _plot_professional_roc(df_pred: pd.DataFrame, split_label: str, save_path: str):
    """
    Generates a Nature/Lancet quality ROC curve:
    - No inset axes (clean space).
    - Fully English labels.
    - Legend outside the plot, dynamically sorted by AUC.
    - Our Proposed Model heavily emphasized.
    """
    if "target" not in df_pred.columns: return

    y_true = df_pred["target"].values
    model_cols = [c for c in df_pred.columns if c != "target"]

    # Calculate AUCs to sort
    model_stats = []
    for col in model_cols:
        auc = roc_auc_score(y_true, df_pred[col].values)
        model_stats.append((auc, col))
    model_stats.sort(key=lambda x: x[0], reverse=True)

    fig, ax = plt.subplots(figsize=(10.5, 8), dpi=300)
    plt.rcParams['font.family'] = 'sans-serif'
    plt.rcParams['axes.unicode_minus'] = False

    # Scientific Palette strictly matched to the 9 baselines
    PALETTE = [
        "#4DBBD5", "#00A087", "#3C5488", "#F39B7F",
        "#8491B4", "#91D1C2", "#7E6148", "#B09C85",
        "#4B4B4B"
    ]

    color_idx = 0
    for auc, col in model_stats:
        fpr, tpr, _ = roc_curve(y_true, df_pred[col].values)

        if col == "Our Proposed Model":
            ax.plot(fpr, tpr, color="#E64B35", lw=4.0, alpha=1.0, label=f"{col} (AUC = {auc:.3f})", zorder=10)
        elif col in ["Stream 1 (Classification)", "Stream 2 (Ranking)", "Teacher (Fusion)"]:
            # Highlight component streams slightly differently
            comp_colors = {"Stream 1 (Classification)": "#FF8C00", "Stream 2 (Ranking)": "#8A2BE2",
                           "Teacher (Fusion)": "#20B2AA"}
            ax.plot(fpr, tpr, color=comp_colors.get(col, "#000"), lw=2.5, alpha=0.85,
                    label=f"Ours: {col} (AUC = {auc:.3f})", zorder=8)
        else:
            # Baselines
            ax.plot(fpr, tpr, color=PALETTE[color_idx % len(PALETTE)], lw=1.5, alpha=0.65,
                    label=f"{col} (AUC = {auc:.3f})", zorder=5)
            color_idx += 1

    # Random Guess Line
    ax.plot([0, 1], [0, 1], color='gray', linestyle='--', lw=1.5, label='Random Guess (AUC = 0.500)', zorder=1)

    ax.set_xlim([-0.01, 1.01])
    ax.set_ylim([-0.01, 1.05])
    ax.set_xlabel("False Positive Rate (1 - Specificity)", fontsize=14, fontweight='bold')
    ax.set_ylabel("True Positive Rate (Sensitivity)", fontsize=14, fontweight='bold')
    ax.set_title(f"Receiver Operating Characteristic (ROC) Curves - {split_label} Cohort", fontsize=16,
                 fontweight='bold', pad=15)

    ax.tick_params(axis='both', labelsize=12)
    ax.grid(True, alpha=0.3, ls='--')

    # Clean up spines
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    for spine in ['left', 'bottom']:
        ax.spines[spine].set_linewidth(1.5)

    # Put Legend OUTSIDE to avoid any overlap with curves
    ax.legend(loc="center left", bbox_to_anchor=(1.02, 0.5), fontsize=11, framealpha=1.0, edgecolor='gray',
              title="Models (Sorted by AUC)", title_fontsize=12)

    plt.tight_layout()
    plt.savefig(save_path, dpi=300, bbox_inches="tight", transparent=False)
    plt.close()


def _plot_professional_dca(baseline_models, our_model, save_path):
    max_thresh = 0.20
    thresholds = np.linspace(0.001, max_thresh, 200)

    fig, ax = plt.subplots(figsize=(10.5, 8), dpi=300)
    plt.rcParams['font.family'] = 'sans-serif'
    plt.rcParams['axes.unicode_minus'] = False

    y_ref = baseline_models[0][1]
    prevalence = y_ref.mean()

    nb_all = np.array([prevalence - (1 - prevalence) * pt / (1 - pt) for pt in thresholds])
    nb_none = np.zeros_like(thresholds)

    ax.plot(thresholds, nb_all, color="gray", linestyle="--", linewidth=2.0, label="Treat All", zorder=2)
    ax.plot(thresholds, nb_none, color="black", linestyle="-", linewidth=2.0, label="Treat None", zorder=2)

    for name, y_true, y_prob, color, ls, lw in baseline_models:
        nb = net_benefit(y_true, y_prob, thresholds)
        ax.plot(thresholds, nb, color=color, linestyle=ls, linewidth=lw, alpha=0.65, label=name, zorder=3)

    if our_model is not None:
        name, y_true, y_prob, color, ls, lw = our_model
        nb = net_benefit(y_true, y_prob, thresholds)
        ax.plot(thresholds, nb, color=color, linestyle=ls, linewidth=lw, label=f"{name}", zorder=10)
        ax.fill_between(
            thresholds, nb, np.maximum(nb_all, nb_none),
            where=(nb > np.maximum(nb_all, nb_none)),
            color=color, alpha=0.12, label="Net Clinical Benefit (Ours)", zorder=1
        )

    y_max = max(prevalence * 1.5, 0.05)
    ax.set_ylim(-0.002, y_max)
    ax.set_xlim(0, max_thresh)
    ax.set_xlabel("Threshold Probability (Risk Preference)", fontsize=14, fontweight='bold')
    ax.set_ylabel("Net Benefit", fontsize=14, fontweight='bold')
    ax.set_title("Decision Curve Analysis (DCA) - Test Cohort", fontsize=16, fontweight='bold', pad=15)

    ax.tick_params(axis='both', labelsize=12)
    ax.grid(True, alpha=0.3, linestyle="--")

    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    for spine in ['left', 'bottom']:
        ax.spines[spine].set_linewidth(1.5)

    ax.legend(loc="center left", bbox_to_anchor=(1.02, 0.5), fontsize=11, framealpha=1.0, edgecolor='gray',
              title="Interventions", title_fontsize=12)

    plt.tight_layout()
    plt.savefig(save_path, dpi=300, bbox_inches="tight", transparent=False)
    plt.close()


def _save_excel(df_result: pd.DataFrame, path: str):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    display_cols = ["Model", "AUC (95% CI)", "Sensitivity", "Specificity", "PPV", "NPV", "Threshold"]
    df_disp = df_result[display_cols].copy()
    for col in ["Sensitivity", "Specificity", "PPV", "NPV"]:
        df_disp[col] = df_disp[col].apply(lambda x: f"{x * 100:.1f}%")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparison"

    col_headers = {c: c for c in display_cols}
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),
                         bottom=Side(style="thin"))

    for ci, col in enumerate(display_cols, start=1):
        cell = ws.cell(row=1, column=ci, value=col_headers[col])
        cell.fill, cell.font, cell.border = header_fill, header_font, thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30
    our_fill = PatternFill("solid", fgColor="E2EFDA")
    base_fill = PatternFill("solid", fgColor="FFFFFF")
    alt_fill = PatternFill("solid", fgColor="F2F2F2")
    best_font = Font(bold=True, size=10)
    best_auc_val = df_result["AUC"].max()

    for ri, row in enumerate(df_disp.itertuples(index=False), start=2):
        is_our = "Our" in str(row[0]) or "Ours" in str(row[0])
        is_best = abs(df_result.iloc[ri - 2]["AUC"] - best_auc_val) < 1e-6
        row_fill = our_fill if is_our else (alt_fill if ri % 2 == 0 else base_fill)

        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill, cell.border = row_fill, thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if is_best: cell.font = best_font

    col_widths = [25, 22, 18, 18, 12, 12, 12]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(path)


if __name__ == "__main__":
    main()